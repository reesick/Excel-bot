"""
Excel Skill — Professional Dashboard Formatter
Pure formatting engine. Receives instruction dict from Claude Analyst,
produces a professional multi-sheet Excel dashboard.

Exposes:
  generate_excel(instruction_dict) → str (output file path)
"""

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# ─── Color Palette ─────────────────────────────────────────────────────────────

PALETTE = {
    "primary":   "1F3864",
    "accent1":   "2E75B6",
    "accent2":   "ED7D31",
    "accent3":   "A9D18E",
    "light_bg":  "EBF3FB",
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "alt_row":   "D6E4F0",
    "white":     "FFFFFF",
    "card_bg":   "F5F5F5",
    "subtle":    "999999",
}


# ─── Style Helpers ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, col_start: int, col_end: int):
    """Apply navy header styling to a row of cells."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=PALETTE["header_fg"], name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor=PALETTE["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws, min_width=10, max_width=35):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ─── Sheet Builders ────────────────────────────────────────────────────────────

def _build_raw_data_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 1: Cleaned raw data as a formatted table."""
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    headers = list(df.columns)
    _style_header_row(ws, 1, 1, len(headers))
    for ci, col in enumerate(headers, 1):
        ws.cell(1, ci, col)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        fill_color = PALETTE["alt_row"] if ri % 2 == 0 else PALETTE["white"]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _autofit_columns(ws)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24


def _build_dashboard_sheet(wb: Workbook, instruction: dict):
    """Sheet 2: Title banner, KPI cards, chart images in 2-column grid."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["accent1"]

    title = instruction.get("dataset_title", "Dashboard")
    dataset_type = instruction.get("dataset_type", "General")

    # Title banner
    ws.row_dimensions[1].height = 50
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"📊  {title}  |  {dataset_type.title()} Dashboard"
    title_cell.font = Font(bold=True, size=20, color=PALETTE["header_fg"], name="Arial")
    title_cell.fill = PatternFill("solid", fgColor=PALETTE["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle / date
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:R2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    sub.font = Font(italic=True, size=10, color=PALETTE["subtle"], name="Arial")
    sub.fill = PatternFill("solid", fgColor="F2F2F2")
    sub.alignment = Alignment(horizontal="right", vertical="center")

    current_row = 4

    # KPI cards
    kpis = instruction.get("kpis", [])
    if kpis:
        ws.row_dimensions[current_row].height = 18
        ws.row_dimensions[current_row + 1].height = 50
        ws.row_dimensions[current_row + 2].height = 30

        col = 2
        for kpi in kpis[:4]:
            # Label row
            ws.merge_cells(start_row=current_row + 1, start_column=col,
                           end_row=current_row + 1, end_column=col + 3)
            label_cell = ws.cell(current_row + 1, col,
                                 f"{kpi.get('icon', '📊')}  {kpi.get('label', '')}")
            label_cell.font = Font(bold=True, size=9, color="888888", name="Arial")
            label_cell.fill = PatternFill("solid", fgColor=PALETTE["card_bg"])
            label_cell.alignment = Alignment(horizontal="center", vertical="bottom")

            # Value row
            ws.merge_cells(start_row=current_row + 2, start_column=col,
                           end_row=current_row + 2, end_column=col + 3)
            val_cell = ws.cell(current_row + 2, col, kpi.get("value", ""))
            val_cell.font = Font(bold=True, size=18, color=PALETTE["accent1"], name="Arial")
            val_cell.fill = PatternFill("solid", fgColor=PALETTE["card_bg"])
            val_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Borders for card effect
            accent_side = Side(style="medium", color=PALETTE["accent1"])
            for r in [current_row + 1, current_row + 2]:
                for c in range(col, col + 4):
                    cell = ws.cell(r, c)
                    cell.border = Border(
                        left=accent_side if c == col else None,
                        right=accent_side if c == col + 3 else None,
                        top=accent_side if r == current_row + 1 else None,
                        bottom=accent_side if r == current_row + 2 else None,
                    )
            col += 5

        current_row += 5

    # Embed charts in 2-column grid
    chart_paths = instruction.get("charts", [])
    col_positions = [2, 11]  # columns B and K
    chart_height_rows = 22

    for idx, cpath in enumerate(chart_paths):
        if not os.path.exists(cpath):
            continue
        col_offset = col_positions[idx % 2]
        row_offset = current_row + (idx // 2) * (chart_height_rows + 2)
        try:
            img = XLImage(cpath)
            img.width = 520
            img.height = 260
            cell_addr = f"{get_column_letter(col_offset)}{row_offset}"
            ws.add_image(img, cell_addr)
        except Exception:
            logger.warning("Failed to embed chart: %s", cpath)


def _build_pivot_sheet(wb: Workbook, df: pd.DataFrame, pivot_instructions: list):
    """Sheet 3: Pivot tables defined by Claude."""
    ws = wb.create_sheet("Pivot Tables")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["accent3"]

    row = 1

    for piv in pivot_instructions:
        title = piv.get("title", "Pivot Table")
        index_col = piv.get("index_col")
        value_col = piv.get("value_col")
        agg = piv.get("agg", "sum")

        if not index_col or not value_col:
            continue
        if index_col not in df.columns or value_col not in df.columns:
            # Try case-insensitive match
            idx_match = next((c for c in df.columns if c.lower() == index_col.lower()), None)
            val_match = next((c for c in df.columns if c.lower() == value_col.lower()), None)
            if not idx_match or not val_match:
                continue
            index_col, value_col = idx_match, val_match

        try:
            pivot_df = df.groupby(index_col)[value_col].agg(agg).reset_index()
            pivot_df.columns = [index_col, f"{agg.title()} of {value_col}"]
            pivot_df = pivot_df.sort_values(pivot_df.columns[1], ascending=False)
        except Exception:
            continue

        # Section title
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=len(pivot_df.columns) + 1)
        t = ws.cell(row, 1, f"📊 {title}")
        t.font = Font(bold=True, size=12, color=PALETTE["header_fg"], name="Arial")
        t.fill = PatternFill("solid", fgColor=PALETTE["accent1"])
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1

        # Headers
        cols = list(pivot_df.columns)
        _style_header_row(ws, row, 1, len(cols))
        for ci, col in enumerate(cols, 1):
            ws.cell(row, ci, col)
        row += 1

        # Data rows
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ri, (_, data_row) in enumerate(pivot_df.iterrows()):
            fill_color = PALETTE["alt_row"] if ri % 2 == 0 else PALETTE["white"]
            for ci, val in enumerate(data_row, 1):
                cell = ws.cell(row, ci, round(val, 2) if isinstance(val, float) else val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = border
                if ci > 1 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
            row += 1

        row += 2  # Gap between pivots

    _autofit_columns(ws)


def _build_insights_sheet(wb: Workbook, insights: list, dataset_title: str):
    """Sheet 4: Numbered insight cards in clean typography."""
    ws = wb.create_sheet("Insights")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["accent2"]

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 80

    # Title
    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = "📝  Executive Insights & Key Findings"
    t.font = Font(bold=True, size=16, color=PALETTE["header_fg"], name="Arial")
    t.fill = PatternFill("solid", fgColor=PALETTE["header_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells("A2:M2")
    s = ws["A2"]
    s.value = dataset_title
    s.font = Font(italic=True, size=11, color=PALETTE["subtle"], name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    row = 4
    for i, insight in enumerate(insights, 1):
        # Number badge
        num_cell = ws.cell(row, 2, str(i))
        num_cell.font = Font(bold=True, size=12, color=PALETTE["header_fg"], name="Arial")
        num_cell.fill = PatternFill("solid", fgColor=PALETTE["accent1"])
        num_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Insight text
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=14)
        text_cell = ws.cell(row, 3, insight)
        text_cell.font = Font(name="Arial", size=11)
        text_cell.alignment = Alignment(wrap_text=True, vertical="center")
        if i % 2 == 0:
            text_cell.fill = PatternFill("solid", fgColor="F7FBFF")

        ws.row_dimensions[row].height = 36
        row += 2  # Gap between insights


# ─── Public API ────────────────────────────────────────────────────────────────

def generate_excel(instruction: dict) -> str:
    """
    Generate a professional multi-sheet Excel dashboard.

    Args:
        instruction: Dict with keys:
            - file_path (str): Path to cleaned CSV
            - charts (list[str]): Paths to chart PNG files
            - insights (list[str]): Plain English insight strings
            - kpis (list[dict]): KPI cards [{label, value, icon}]
            - pivot_instructions (list[dict]): Pivot definitions
            - dataset_title (str): Report title
            - dataset_type (str): Data category
            - output_path (str): Where to save the xlsx

    Returns:
        Path to the generated .xlsx file
    """
    file_path = instruction["file_path"]
    output_path = instruction.get("output_path")

    if not output_path:
        base = os.path.splitext(os.path.basename(file_path))[0]
        output_path = os.path.join(os.path.dirname(file_path), f"{base}_dashboard.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Load cleaned data
    try:
        df = pd.read_csv(file_path)
    except Exception:
        df = pd.read_excel(file_path)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _build_raw_data_sheet(wb, df)
    _build_dashboard_sheet(wb, instruction)

    pivot_instructions = instruction.get("pivot_instructions", [])
    if pivot_instructions:
        _build_pivot_sheet(wb, df, pivot_instructions)

    insights = instruction.get("insights", [])
    if insights:
        _build_insights_sheet(wb, insights, instruction.get("dataset_title", "Report"))

    # Order sheets
    desired = ["Raw Data", "Dashboard", "Pivot Tables", "Insights"]
    wb._sheets = [wb[s] for s in desired if s in wb.sheetnames]

    wb.save(output_path)
    logger.info("Excel dashboard saved: %s", output_path)
    return output_path
